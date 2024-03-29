import os
from typing import re
from bs4 import BeautifulSoup
from time import sleep
import requests
import urllib3
import ssl
from openpyxl import load_workbook, Workbook
import pandas as pd
import re
from datetime import datetime


class CustomHttpAdapter(requests.adapters.HTTPAdapter):
    # "Transport adapter" that allows us to use custom ssl_context.

    def __init__(self, ssl_context=None, **kwargs):
        self.ssl_context = ssl_context
        super().__init__(**kwargs)

    def init_poolmanager(self, connections, maxsize, block=False):
        self.poolmanager = urllib3.poolmanager.PoolManager(
            num_pools=connections, maxsize=maxsize,
            block=block, ssl_context=self.ssl_context)


# Reading a list of elements from a source file.
class Exel_RW:

    def read_exel(file_name):
        workbook = pd.read_excel(file_name)
        list_product = []
        for elements in workbook.values:
            elements = list(elements)
            elements2 = []
            for i in elements:
                if type(i) == str:
                    i = i.replace('\xa0', ' ')
                elements2.append(i)
            list_product.append(elements2)
        return list_product

    def write_exel(write_lists, file_name, sheet_name=0):
        if file_name not in os.listdir():
            workbook = Workbook()
            workbook.save(file_name)
            workbook.close()
        workbook = load_workbook(file_name, data_only=True)
        if sheet_name == 0:
            worksheet = workbook[workbook.sheetnames[0]]
        elif sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[sheet_name]
        for list_values in write_lists:
            worksheet.append(list_values)
        workbook.save(file_name)
        workbook.close()


class Analysis:

    def __init__(self, list_data, list_original):
        self.list_data = list_data
        self.list_original = list_original

    def price_check(self):
        price_data = int(self.list_data[2])
        price_original = int(self.list_original[0])
        if price_data < price_original:
            percentage_value = int(abs((100 - price_data / price_original * 100)))
            if percentage_value > 50:
                return False
            else:
                return f"-{percentage_value}"
        elif price_data > price_original:
            percentage_value = int(abs((price_data / price_original) * 100 - 100))
            if percentage_value > 30:
                return False
            else:
                return str(percentage_value)
        else:
            return 0



def duplicate_list_exception(list_product, lists_products):
    lists_products_complete = []
    for product_list in lists_products:
        if product_list != list_product:
            lists_products_complete.append(product_list)
    return lists_products_complete


def get_lists_original_product(dict_product, original):
    lists_original_products = []
    counter_analog = 0
    for dict_analog in dict_product:
        if counter_analog != 7:
            if dict_analog["quantity"] == 1000:
                dict_analog["quantity"] = "под заказ"
            list_product = [str(dict_analog['price']),
                            str(dict_analog["rating"]),
                            str(dict_analog["quantity"]),
                            str(dict_analog["delivery"]),
                           ]
            if int(list_product[3]) < 31 and list_product != original:
                lists_original_products.append(list_product)
                counter_analog += 1
    return lists_original_products


def get_legacy_session():
    ctx = ssl.create_default_context(ssl.Purpose.SERVER_AUTH)
    ctx.options |= 0x4  # OP_LEGACY_SERVER_CONNECT
    session = requests.session()
    session.mount('https://', CustomHttpAdapter(ctx))
    return session


def get_html(url):
    try:
        html = get_legacy_session().get(url, timeout=10, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:45.0) Gecko/20100101 Firefox/45.0'})
    except:
        sleep(2)
        html = get_html(url)
    return html


def get_url_product_emex(vendor_cod):
    dict_product = get_emex_dict_products(vendor_cod)
    try:
        return dict_product["makes"]["list"][0]["url"]
    except:
        return False


def get_emex_original_list_product(vendor_cod):

    list_product = []
    url = get_url_product_emex(vendor_cod)
    if url:
        url = "https://emex.ru/products/" + url
        html_product = get_html(url).text
        parser = BeautifulSoup(html_product, "lxml")

        availability = parser.find(class_="sc-b0f3936c-1 kHZHVQ")

        if availability != None:
            regex_num = re.compile('\d+')
            rating = parser.find(class_="sc-b0f3936c-1 kHZHVQ").text
            try:
                quantity = "".join(regex_num.findall(parser.find(class_="sc-d67ce909-11 sc-d67ce909-13 fuNkfc csqgZG").text))
            except:
                quantity = "под заказ"
            delivery = "".join(regex_num.findall(parser.find(class_="sc-d67ce909-11 sc-d67ce909-14 fuNkfc jtgcED").text))
            price = "".join(regex_num.findall(parser.find(class_="sc-d67ce909-11 sc-d67ce909-15 fuNkfc gXBVKh").text))
            list_product.append(price)
            list_product.append(rating)
            list_product.append(quantity)
            list_product.append(delivery)
        else:
            return False
        return [list_product]
    return False


def get_lists_dict_originals_or_analogs(dict_product, list_name):
    lists_dict_analogs_completed = []
    if  "analogs" in dict_product:
        list_dict_analogs = dict_product[list_name]
        for analog_dict in list_dict_analogs:
            offers = analog_dict['offers']
            for offer in offers:
                lists_dict_analogs_completed.append({
                    "vendor_cod": analog_dict["detailNum"],
                    "make": analog_dict['make'],
                    "name": analog_dict['name'],
                    "price": offer['displayPrice']['value'],
                    "rating": offer['rating2']['rating'],
                    "quantity": offer['quantity'],
                    "delivery": offer['delivery']['value']
                })
    return lists_dict_analogs_completed


def get_lists_product(input_lists):
    write_list = []
    write_product_dict = {}


    count = len(input_lists)
    for list_product in input_lists:
        print(list_product)#########################

        print(count)################################
        count -= 1##############################################


        list_original_product = list_product

        vendor_cod = str(list_product[-1])



        if vendor_cod not in ["nan", "-"]:

            emex_list_original_product = get_emex_original_list_product(vendor_cod)

            if emex_list_original_product:         # если продукт найден
                #list_original_product += emex_list_original_product

                print(vendor_cod)###################################################


                dict_product = get_emex_dict_products(vendor_cod)
                lists_dict_originals = get_lists_dict_originals_or_analogs(dict_product, "originals")
                lists_original_products_emex = get_lists_original_product(lists_dict_originals, emex_list_original_product)
                lists_dict_analogs = get_lists_dict_originals_or_analogs(dict_product, "analogs")

                emex_list_original_product += duplicate_list_exception(emex_list_original_product[0], lists_original_products_emex)



                counter_analog = 0
                counter_product = 0
                for product_list in emex_list_original_product:

                    #write_product_dict[list_original_product[6]] = {}
                    #write_product_dict[list_original_product[6]][counter_product] = list_original_product + product_list
                    write_list_product = list_original_product + product_list


                    for dict_analog in lists_dict_analogs:
                        if counter_analog != 5:
                            if dict_analog["quantity"] == 1000:
                                dict_analog["quantity"] = "под заказ"
                            list_analog = [dict_analog['make'],
                                           dict_analog['vendor_cod'],
                                           dict_analog['price'],
                                           f"https://emex.ru/products/{dict_analog['vendor_cod']}/{dict_analog['make']}/29241",
                                           dict_analog["rating"],
                                           dict_analog["quantity"],
                                           dict_analog["delivery"],
                                           ]
                            check_by_criterion = Analysis(list_analog, product_list).price_check()
                            if check_by_criterion:
                                counter_analog += 1

                                write_list_product += list_analog + [check_by_criterion]
                                #write_product_dict[list_original_product[6]][counter_product] += list_analog + [check_by_criterion]
                        else:
                            counter_analog = 0
                            break
                    write_list.append(write_list_product)
                counter_product += 1
            else:
                #write_product_dict[list_original_product[6]] = {}
                #write_product_dict[list_original_product[6]][0] = list_original_product + [0]
                write_list.append(list_product + [0])

    return write_list


def write_list_data(lists_product):
    column_names = [["ID",
                     "Марка ТС",
                     "Модель ТС",
                     "Тип кузова",
                     "Ценовой сегмент",
                     "Наименование запчасти",
                     "Артикул (оригинал)",
                     "Цена EMEX на артикул оригинала",
                     "Рейтинг",
                     "Наличие в шт.",
                     "Количество дней доставки",
                     "Бренд Аналога  1",
                     "Артикул Аналога 1",
                     "Цена Аналога 1",
                     "Источник цен от emex (ссылка)",
                     "Рейтинг",
                     "Наличие в шт.",
                     "Количество дней доставки",
                     "% разница стоимости",
                     "Бренд Аналога  2",
                     "Артикул Аналога 2",
                     "Цена Аналога 2",
                     "Источник цен от emex (ссылка)",
                     "Рейтинг",
                     "Наличие в шт.",
                     "Количество дней доставки",
                     "% разница стоимости",
                     "Бренд Аналога  3",
                     "Артикул Аналога 3",
                     "Цена Аналога 3",
                     "Источник цен от emex (ссылка)",
                     "Рейтинг",
                     "Наличие в шт.",
                     "Количество дней доставки",
                     "% разница стоимости",
                     "Бренд Аналога  4",
                     "Артикул Аналога 4",
                     "Цена Аналога 4",
                     "Источник цен от emex (ссылка)",
                     "Рейтинг",
                     "Наличие в шт.",
                     "Количество дней доставки",
                     "% разница стоимости",
                     "Бренд Аналога  5",
                     "Артикул Аналога 5",
                     "Цена Аналога 5",
                     "Источник цен от emex (ссылка)",
                     "Рейтинг",
                     "Наличие в шт.",
                     "Количество дней доставки",
                     "% разница стоимости",]]
    return column_names + lists_product


def get_emex_dict_products(vendor_cod):
    url_part1 = "https://emex.ru/api/search/search?detailNum="
    url_part2 = "&locationId=29241&showAll=true"
    url = url_part1 + vendor_cod + url_part2
    dict_product = (get_html(url).json().get('searchResult'))
    return dict_product


def main():
    input_list = Exel_RW.read_exel("input.xlsx")
    product_lists = get_lists_product(input_list)
    if "korzina.xlsx" not in os.listdir():
        product_lists = write_list_data(product_lists)
    Exel_RW.write_exel(product_lists, "korzina.xlsx")


if __name__ == '__main__':
    main()
    pass
